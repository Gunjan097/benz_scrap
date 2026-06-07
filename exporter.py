import os
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_STUDENT_COLS = [
    "objectId", "qr", "photono", "grno", "class", "dob",
    "studentname", "fathername", "mothername", "address",
    "phone1", "phone2", "createdAt", "updatedAt",
]
_PHOTO_COLS   = ["grno", "studentname", "photono", "photo_name", "local_file"]
_FILL         = PatternFill("solid", fgColor="1F4E79")
_FONT         = Font(bold=True, color="FFFFFF")
_ALIGN        = Alignment(horizontal="center", vertical="center")


def export(records: list, xlsx_path: str, progress_cb=None) -> int:
    """
    Writes two sheets: Students and Photos.
    Always downloads photos into a 'photos' folder next to the Excel file.
    Returns number of photos downloaded successfully.
    """
    photo_dir = os.path.join(os.path.dirname(os.path.abspath(xlsx_path)), "photos")
    os.makedirs(photo_dir, exist_ok=True)

    wb = openpyxl.Workbook()
    _write_sheet(wb.active, "Students", records, _STUDENT_COLS,
                 lambda r, c: r.get(c, "") or "")

    photo_rows, downloaded = _build_photo_rows(records, photo_dir, progress_cb)
    _write_sheet(wb.create_sheet("Photos"), "Photos", photo_rows, _PHOTO_COLS,
                 lambda r, c: r.get(c, "") or "")

    wb.save(xlsx_path)
    return downloaded


def _build_photo_rows(records: list, photo_dir: str, progress_cb=None):
    rows, downloaded = [], 0
    total = len(records)
    for i, r in enumerate(records):
        photo = r.get("photo") or {}
        url   = photo.get("url", "")
        name  = photo.get("name", "")
        local = _download_photo(url, name, photo_dir) if url else ""
        if local:
            downloaded += 1
        rows.append({
            "grno":        r.get("grno", ""),
            "studentname": r.get("studentname", ""),
            "photono":     r.get("photono", ""),
            "photo_name":  name,
            "local_file":  local,
        })
        if progress_cb:
            progress_cb(i + 1, total)
    return rows, downloaded


def _download_photo(url: str, name: str, photo_dir: str) -> str:
    try:
        filename = name or url.split("/")[-1]
        dest = os.path.join(photo_dir, filename)
        if os.path.exists(dest):
            return dest
        resp = requests.get(url, timeout=20)
        if resp.status_code == 200:
            with open(dest, "wb") as f:
                f.write(resp.content)
            return dest
    except Exception:
        pass
    return ""


def _write_sheet(ws, title: str, records: list, cols: list, getter):
    ws.title = title
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill, cell.font, cell.alignment = _FILL, _FONT, _ALIGN

    for ri, rec in enumerate(records, 2):
        for ci, col in enumerate(cols, 1):
            ws.cell(row=ri, column=ci, value=str(getter(rec, col)))

    for ci, col in enumerate(cols, 1):
        width = max(len(col), max(
            (len(str(ws.cell(r, ci).value or "")) for r in range(2, ws.max_row + 1)),
            default=0,
        ))
        ws.column_dimensions[get_column_letter(ci)].width = min(width + 2, 45)
